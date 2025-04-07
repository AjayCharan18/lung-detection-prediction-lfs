# Set matplotlib backend first to avoid Tkinter warnings
import matplotlib
matplotlib.use('Agg')  # Use non-interactive backend
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from sklearn.model_selection import train_test_split, RandomizedSearchCV, StratifiedKFold
from sklearn.preprocessing import StandardScaler, OneHotEncoder
from sklearn.compose import ColumnTransformer
from sklearn.pipeline import Pipeline
from sklearn.ensemble import RandomForestClassifier
from sklearn.metrics import (roc_auc_score, balanced_accuracy_score, 
                           f1_score, classification_report, confusion_matrix)
from imblearn.ensemble import BalancedRandomForestClassifier
from imblearn.over_sampling import SMOTE
from imblearn.pipeline import Pipeline as ImbPipeline
from sklearn.impute import SimpleImputer
import joblib
import warnings
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as ExcelImage
import os
from datetime import datetime
warnings.filterwarnings('ignore')

def main():
    # Load and prepare data
    try:
        df = pd.read_csv('dataset_med.csv')
        print(f"Dataset loaded successfully with shape: {df.shape}")
        
        # Downsample for development (comment out for full run)
        # df = df.sample(frac=0.1, random_state=42)
        # print(f"Using subset of data with shape: {df.shape}")
        
    except FileNotFoundError:
        print("Error: File 'dataset_med.csv' not found.")
        return
    except Exception as e:
        print(f"Error loading dataset: {str(e)}")
        return

    # Data preprocessing function
    def preprocess_data(df, is_training=True):
        try:
            # Create categorical features
            df['age_group'] = pd.cut(df['age'], bins=[0,40,60,80,100], 
                                    labels=['0-40','41-60','61-80','80+'])
            df['bmi_category'] = pd.cut(df['bmi'], bins=[0,18.5,25,30,100], 
                                      labels=['Underweight','Normal','Overweight','Obese'])
            
            # Ensure categorical features are of type category for training, object for prediction
            if is_training:
                df['age_group'] = df['age_group'].astype('category')
                df['bmi_category'] = df['bmi_category'].astype('category')
            else:
                df['age_group'] = df['age_group'].astype('object')
                df['bmi_category'] = df['bmi_category'].astype('object')
            
            # Drop problematic columns that might cause leakage
            cols_to_drop = ['id', 'diagnosis_date', 'end_treatment_date', 'country']
            df = df.drop([col for col in cols_to_drop if col in df.columns], axis=1)
            
            return df
        except Exception as e:
            print(f"Error in preprocessing: {str(e)}")
            return None

    df_processed = preprocess_data(df.copy(), is_training=True)
    if df_processed is None:
        return

    # Define features and target
    X = df_processed.drop('survived', axis=1)
    y = df_processed['survived']

    # Split data with stratification
    X_train, X_test, y_train, y_test = train_test_split(
        X, y, test_size=0.2, random_state=42, stratify=y)

    # Define preprocessing
    numerical_features = ['age', 'bmi', 'cholesterol_level']
    categorical_features = ['gender', 'cancer_stage', 'family_history', 'smoking_status',
                           'hypertension', 'asthma', 'cirrhosis', 'other_cancer', 
                           'treatment_type', 'age_group', 'bmi_category']

    preprocessor = ColumnTransformer(
        transformers=[
            ('num', Pipeline([
                ('imputer', SimpleImputer(strategy='median')),
                ('scaler', StandardScaler())
            ]), numerical_features),
            ('cat', OneHotEncoder(handle_unknown='ignore'), categorical_features)
        ])

    # Define modeling approaches with improved parameters
    approaches = {
        # Approach 1: Balanced Random Forest
        'Balanced_RF': {
            'model': Pipeline([
                ('preprocessor', preprocessor),
                ('classifier', BalancedRandomForestClassifier(
                    random_state=42,
                    sampling_strategy='auto',
                    n_jobs=-1,  # Use all available cores
                    class_weight='balanced_subsample'
                ))
            ]),
            'params': {
                'classifier__n_estimators': [100, 200, 300],
                'classifier__max_depth': [5, 10, 20, None],
                'classifier__min_samples_leaf': [1, 2, 4],
                'classifier__max_features': ['sqrt', 'log2', 0.5]
            }
        },
        
        # Approach 2: SMOTE + Random Forest
        'SMOTE_RF': {
            'model': ImbPipeline([
                ('preprocessor', preprocessor),
                ('sampling', SMOTE(random_state=42, sampling_strategy=0.5)),
                ('classifier', RandomForestClassifier(
                    random_state=42,
                    class_weight='balanced',
                    n_jobs=-1  # Use all available cores
                ))
            ]),
            'params': {
                'classifier__n_estimators': [100, 200, 300],
                'classifier__max_depth': [5, 10, 20, None],
                'classifier__min_samples_split': [2, 5, 10],
                'sampling__k_neighbors': [3, 5, 7]
            }
        }
    }

    # Create Excel workbook with timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_filename = f'lung_cancer_prediction_results_{timestamp}.xlsx'
    wb = Workbook()
    ws_results = wb.active
    ws_results.title = "Model Results"
    ws_feature_importance = wb.create_sheet("Feature Importance")
    ws_predictions = wb.create_sheet("Example Predictions")

    # Model evaluation with better CV strategy
    results = {}
    best_score = 0
    best_model = None

    print("\n=== Evaluating Different Approaches ===")
    for approach_name, approach in approaches.items():
        print(f"\nEvaluating {approach_name}...")
        
        try:
            search = RandomizedSearchCV(
                approach['model'],
                approach['params'],
                cv=StratifiedKFold(n_splits=5, shuffle=True, random_state=42),
                scoring='roc_auc',
                n_jobs=1,  # Set to 1 to avoid memory issues
                verbose=3,
                n_iter=15,  # Increased from 10 to 15 for better parameter search
                refit=True,
                random_state=42
            )
            
            search.fit(X_train, y_train)
            
            # Evaluate best model
            current_model = search.best_estimator_
            y_pred = current_model.predict(X_test)
            y_proba = current_model.predict_proba(X_test)[:, 1]
            
            # Calculate metrics
            metrics = {
                'roc_auc': roc_auc_score(y_test, y_proba),
                'balanced_accuracy': balanced_accuracy_score(y_test, y_pred),
                'f1_score': f1_score(y_test, y_pred),
                'best_params': str(search.best_params_),  # Convert to string for Excel
                'model': approach_name
            }
            
            # Print classification report
            print(f"\nClassification Report for {approach_name}:")
            print(classification_report(y_test, y_pred))
            
            # Plot and save confusion matrix
            plt.figure(figsize=(6,4))
            sns.heatmap(confusion_matrix(y_test, y_pred), annot=True, fmt='d', cmap='Blues',
                        xticklabels=['Not Survived', 'Survived'],
                        yticklabels=['Not Survived', 'Survived'])
            plt.title(f'Confusion Matrix - {approach_name}')
            plt.xlabel('Predicted')
            plt.ylabel('Actual')
            conf_matrix_path = f'confusion_matrix_{approach_name}.png'
            plt.savefig(conf_matrix_path, bbox_inches='tight', dpi=300)
            plt.close()
            
            results[approach_name] = metrics
            print(f"{approach_name} - ROC AUC: {metrics['roc_auc']:.4f}, F1: {metrics['f1_score']:.4f}")
            
            if metrics['roc_auc'] > best_score:
                best_score = metrics['roc_auc']
                best_model = current_model
                
        except Exception as e:
            print(f"Error in {approach_name}: {str(e)}")
            continue

    # Results summary
    print("\n=== Final Results ===")
    results_df = pd.DataFrame.from_dict(results, orient='index')
    print(results_df[['roc_auc', 'balanced_accuracy', 'f1_score']].sort_values('roc_auc', ascending=False))

    # Write results to Excel
    ws_results.append(["Model Comparison Results"])
    ws_results.append(["Model", "ROC AUC", "Balanced Accuracy", "F1 Score", "Best Parameters"])
    for model_name, metrics in results.items():
        ws_results.append([
            model_name,
            metrics['roc_auc'],
            metrics['balanced_accuracy'],
            metrics['f1_score'],
            metrics['best_params']
        ])

    # Feature Importance from best model
    if best_model is not None:
        best_approach = max(results.items(), key=lambda x: x[1]['roc_auc'])[0]
        print(f"\nBest approach: {best_approach} (ROC AUC: {best_score:.4f})")
        
        try:
            # Get feature names
            num_features = numerical_features.copy()
            cat_transformer = best_model.named_steps['preprocessor'].named_transformers_['cat']
            cat_features = cat_transformer.get_feature_names_out(categorical_features)
            all_features = np.concatenate([num_features, cat_features])
            
            # Get importances
            classifier = best_model.named_steps['classifier']
            if hasattr(classifier, 'feature_importances_'):
                importances = classifier.feature_importances_
            elif hasattr(classifier, 'estimators_'):
                importances = np.mean([est.feature_importances_ for est in classifier.estimators_], axis=0)
            else:
                raise AttributeError("Classifier doesn't have feature importances")
            
            # Create importance dataframe
            feature_importance = pd.DataFrame({
                'feature': all_features,
                'importance': importances
            }).sort_values('importance', ascending=False)
            
            # Write feature importance to Excel
            ws_feature_importance.append(["Feature Importance Analysis"])
            for r in dataframe_to_rows(feature_importance, index=False, header=True):
                ws_feature_importance.append(r)
            
            # Plot and save top 20 features
            plt.figure(figsize=(12,8))
            sns.barplot(x='importance', y='feature', 
                        data=feature_importance.head(20),
                        palette='viridis')
            plt.title(f'Top 20 Features - {best_approach}')
            plt.xlabel('Feature Importance Score')
            plt.ylabel('Feature Name')
            plt.tight_layout()
            feature_importance_path = 'feature_importance.png'
            plt.savefig(feature_importance_path, bbox_inches='tight', dpi=300)
            plt.close()
            
            # Add image to Excel
            img = ExcelImage(feature_importance_path)
            ws_feature_importance.add_image(img, 'D2')
            
        except Exception as e:
            print(f"Could not plot feature importances: {str(e)}")

        # Save the best model with timestamp
        model_filename = f'best_lung_cancer_model_{timestamp}.pkl'
        joblib.dump(best_model, model_filename)
        print(f"\nBest model saved as '{model_filename}'")
    else:
        print("\nNo successful models were trained.")

    # Enhanced prediction function
    def predict_survival(patient_data, model_path=None):
        """
        Predict survival probability with clear, formatted output
        Args:
            patient_data: Dictionary containing patient features
            model_path: Path to saved model (defaults to most recent)
        Returns:
            Dictionary with prediction results and status
        """
        try:
            if model_path is None:
                # Find the most recent model
                model_files = [f for f in os.listdir() if f.startswith('best_lung_cancer_model_')]
                if not model_files:
                    return {"error": "No trained model found", "status": "error"}
                model_path = sorted(model_files)[-1]  # Get most recent
            
            # Load model
            model = joblib.load(model_path)
            
            # Check for required numerical features
            missing_features = [f for f in numerical_features if f not in patient_data]
            if missing_features:
                return {
                    "error": f"Missing numerical features: {', '.join(missing_features)}",
                    "status": "error"
                }
            
            # Prepare input data
            patient_df = pd.DataFrame([patient_data])
            patient_df = preprocess_data(patient_df, is_training=False)
            if patient_df is None:
                return {"error": "Data preprocessing failed", "status": "error"}
            
            # Make prediction
            proba = model.predict_proba(patient_df)[0, 1]
            prediction = "Yes" if proba >= 0.5 else "No"
            confidence = "high" if abs(proba - 0.5) > 0.3 else "moderate" if abs(proba - 0.5) > 0.15 else "low"
            
            # Get feature importances
            num_features = numerical_features.copy()
            cat_transformer = model.named_steps['preprocessor'].named_transformers_['cat']
            cat_features = cat_transformer.get_feature_names_out(categorical_features)
            all_features = np.concatenate([num_features, cat_features])
            
            classifier = model.named_steps['classifier']
            if hasattr(classifier, 'feature_importances_'):
                importances = classifier.feature_importances_
            elif hasattr(classifier, 'estimators_'):
                importances = np.mean([est.feature_importances_ for est in classifier.estimators_], axis=0)
            
            X_processed = model.named_steps['preprocessor'].transform(patient_df)
            
            # Create contribution dataframe
            contributions = pd.DataFrame({
                'feature': all_features,
                'importance': importances,
                'value': X_processed[0]
            }).sort_values('importance', ascending=False)
            
            # Get top 3 contributing factors
            top_factors = []
            for _, row in contributions.head(3).iterrows():
                if row['feature'] in numerical_features:
                    factor_name = row['feature'].replace('_', ' ').title()
                    if factor_name == 'Bmi':
                        factor_name = 'BMI'
                    factor_value = f"{float(patient_data[row['feature']]):.1f}"
                    top_factors.append(f"{factor_name}: {factor_value}")
                elif row['value'] == 1:
                    feature_parts = row['feature'].split('_')
                    factor_name = feature_parts[0].title()
                    factor_value = ' '.join(feature_parts[1:]).title()
                    top_factors.append(f"{factor_name}: {factor_value}")
            
            return {
                "prediction": prediction,
                "probability": float(proba),
                "confidence": confidence,
                "top_factors": top_factors,
                "status": "success",
                "model_used": model_path
            }
        
        except Exception as e:
            return {
                "error": str(e),
                "status": "error"
            }

    # Example usage and Excel output
    if best_model is not None:
        print("\n=== Example Prediction ===")
        example_patient = {
            'age': 60,
            'gender': 'Female',
            'cancer_stage': 'Stage II',
            'family_history': 'No',
            'smoking_status': 'Former smoker',
            'bmi': 24,
            'cholesterol_level': 180,
            'hypertension': 0,
            'asthma': 0,
            'cirrhosis': 0,
            'other_cancer': 0,
            'treatment_type': 'Surgery'
        }

        prediction = predict_survival(example_patient)
        
        if prediction.get('status') == 'success':
            print("\nPrediction Results:")
            print(f"Likely to Survive: {prediction['prediction']}")
            print(f"Confidence: {prediction['confidence']} ({prediction['probability']:.1%})")
            print(f"Model Used: {prediction['model_used']}")
            print("\nTop Contributing Factors:")
            for i, factor in enumerate(prediction['top_factors'], 1):
                print(f"{i}. {factor}")
            
            # Write prediction to Excel
            ws_predictions.append(["Example Prediction Results"])
            ws_predictions.append(["Likely to Survive:", prediction['prediction']])
            ws_predictions.append(["Probability:", f"{prediction['probability']:.1%}"])
            ws_predictions.append(["Confidence:", prediction['confidence']])
            ws_predictions.append(["Model Used:", prediction['model_used']])
            ws_predictions.append([])
            ws_predictions.append(["Top Contributing Factors:"])
            for i, factor in enumerate(prediction['top_factors'], 1):
                ws_predictions.append([f"{i}.", factor])
        else:
            print(f"\nPrediction failed: {prediction.get('error', 'Unknown error')}")
            ws_predictions.append(["Prediction failed:", prediction.get('error', 'Unknown error')])

    # Save the Excel file
    try:
        wb.save(excel_filename)
        print(f"\nAll results saved to {excel_filename}")
    except Exception as e:
        print(f"\nError saving Excel file: {str(e)}")

    # Clean up temporary image files
    for file in os.listdir():
        if file.startswith('confusion_matrix_') and file.endswith('.png'):
            os.remove(file)
    if os.path.exists('feature_importance.png'):
        os.remove('feature_importance.png')

if __name__ == "__main__":
    main()